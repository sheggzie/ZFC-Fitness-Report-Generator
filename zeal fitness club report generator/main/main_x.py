from openpyxl import load_workbook
from PIL import Image, ImageDraw, ImageFont

# Load the Excel workbook
workbook = load_workbook('april.xlsx')

# Select the active worksheet
worksheet = workbook['Actually']

# Load the font you want to use
font = ImageFont.load_default()

# Create an empty image
img = Image.new('RGB', (600, 500), color = (255, 255, 255))

# Initialize ImageDraw object
draw = ImageDraw.Draw(img)

# Set initial position for text drawing
x = 10
y = 10

text = workbook.active

# draw.text((x, y), text, fill=(0, 0, 0), font=font)

# # Loop through rows and columns to extract text and draw onto image
# for row in worksheet.iter_rows():  
#     for cell in row:
#         # Extract text from cell
#         text = str(cell.value)
#         # Draw text onto image
#         draw.text((x, y), text, fill=(0, 0, 0), font=font)
#         # Move to the next position
#         y += 20  # Adjust this value based on your font size or layout
#     y += 10  # Add some space between rows

# Save the image
img.save('output_image_new.png')

print("Image created successfully!")
